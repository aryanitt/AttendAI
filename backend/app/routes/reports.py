import io
from calendar import monthrange
from datetime import datetime, timezone

from bson import ObjectId
from flask import Blueprint, jsonify, request, send_file
from openpyxl import Workbook

from ..auth_utils import require_auth
from ..db import get_db
from ..utils import parse_oid

bp = Blueprint("reports", __name__)


def _owned_class(db, class_id: str, teacher_id: ObjectId):
    try:
        cid = parse_oid(class_id)
    except ValueError:
        return None, None
    c = db.classes.find_one({"_id": cid, "teacher_id": teacher_id})
    return c, cid


def _xlsx_response(data: bytes, filename: str):
    return send_file(
        io.BytesIO(data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@bp.get("/classes/<class_id>/reports/daily")
def report_daily(class_id):
    db = get_db()
    # Mock teacher_id for debug
    teacher_id = ObjectId("507f1f77bcf86cd799439011")
    c, cid = _owned_class(db, class_id, teacher_id)
    if not c:
        return jsonify({"error": "Class not found"}), 404
    day = request.args.get("date") or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    rows = list(db.attendance.find({"class_id": cid, "date": day}))
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily"
    ws.append(["Date", "Time", "Roll", "Name", "Status"])
    for r in rows:
        st = db.students.find_one({"_id": r["student_id"]})
        ws.append(
            [
                r.get("date"),
                r.get("time"),
                st["roll_number"] if st else "",
                st["name"] if st else "",
                r.get("status"),
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return _xlsx_response(buf.read(), f"attendance_daily_{day}.xlsx")


@bp.get("/classes/<class_id>/reports/monthly")
def report_monthly(class_id):
    db = get_db()
    # Mock teacher_id for debug
    teacher_id = ObjectId("507f1f77bcf86cd799439011")
    c, cid = _owned_class(db, class_id, teacher_id)
    if not c:
        return jsonify({"error": "Class not found"}), 404
    ym = request.args.get("month")  # YYYY-MM
    if not ym or len(ym) != 7:
        ym = datetime.now(timezone.utc).strftime("%Y-%m")
    y, m = int(ym[:4]), int(ym[5:7])
    last = monthrange(y, m)[1]
    prefix = ym
    rows = list(
        db.attendance.find(
            {
                "class_id": cid,
                "date": {"$gte": f"{prefix}-01", "$lte": f"{prefix}-{last:02d}"},
            }
        ).sort("date", 1)
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly"
    ws.append(["Date", "Time", "Roll", "Name", "Status"])
    for r in rows:
        st = db.students.find_one({"_id": r["student_id"]})
        ws.append(
            [
                r.get("date"),
                r.get("time"),
                st["roll_number"] if st else "",
                st["name"] if st else "",
                r.get("status"),
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return _xlsx_response(buf.read(), f"attendance_monthly_{ym}.xlsx")


@bp.get("/classes/<class_id>/reports/student/<student_id>")
def report_student(class_id, student_id):
    db = get_db()
    # Mock teacher_id for debug
    teacher_id = ObjectId("507f1f77bcf86cd799439011")
    c, cid = _owned_class(db, class_id, teacher_id)
    if not c:
        return jsonify({"error": "Class not found"}), 404
    try:
        sid = parse_oid(student_id)
    except ValueError:
        return jsonify({"error": "Invalid student id"}), 400
    st = db.students.find_one({"_id": sid, "class_id": cid})
    if not st:
        return jsonify({"error": "Student not found"}), 404
    rows = list(db.attendance.find({"class_id": cid, "student_id": sid}).sort("date", 1))
    wb = Workbook()
    ws = wb.active
    ws.title = "Student"
    ws.append(["Date", "Time", "Status"])
    for r in rows:
        ws.append([r.get("date"), r.get("time"), r.get("status")])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_roll = "".join(c for c in st["roll_number"] if c.isalnum() or c in "-_")
    return _xlsx_response(
        buf.read(), f"student_{safe_roll or sid}_attendance.xlsx"
    )
